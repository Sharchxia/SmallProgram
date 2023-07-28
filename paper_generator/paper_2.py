import openpyxl as opl
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as msb
import random as rd
import time
from PIL import Image, ImageTk, ImageSequence


class Ques:  # 题目题头，题目选项列坐标数组(或者答案），题目在test.xlsx中的行坐标，此类方便后续随时使用选好的题
    def __init__(self, title, answer, num):  # 对于选择题answer包含打乱后的选项编号（如[2,3,1,4]）说明正确答案是第三个，因为正确答案对应的值恒为1（方便改卷）
        self.title = title  # 题目题干，文字描述
        self.answer = answer
        self.num = num  # 题目被随机选出来后在.xlxs文件中的行号，方便以后查找

    def show(self):
        print(self.title, self.answer, self.num)  # 主要用于调试时的可视化，程序不需要这个


def answer_process(my_str: str) -> str:  # 去除输入答案中的空格，换行符， 保证能与答案相应
    my_str = my_str.replace(' ', '').replace('\n', '')
    return my_str


def generate_question_title(question_type: str, n_of_ques: int) -> list:  # 题目的类型，题目数量
    a, b = 0, 0
    if question_type == 'choose' or question_type == 'gap_filling':
        a, b = 1, 55  # 选择题和填空题所处位置， 可随题库改变做出改变
    elif question_type == 'check':
        a, b = 56, 96  # 判断题所处位置
    title_num = []  # 用于储存随机数产生的题号，以后从.xlxs文件中取对应行号的题
    while len(title_num) < n_of_ques:
        j = rd.randint(a, b)
        if j in title_num:  # 保证题目不会被重复选取
            continue
        else:
            title_num.append(j)
    return title_num


def ques_num(q_type: str, num: int) -> dict or list:  # 题目类型， 题目数量
    ques_nu = generate_question_title(q_type, num)  # 获得题目的题号（对应.xlxs文件的行号）
    if q_type == 'check':  # 对于判断题不需要提取题目的打乱选项
        return ques_nu  # 所以不处理直接返回
    ques_box = {}  # 字典用于题号和选项的一一对应
    while len(ques_box) < len(ques_nu):
        ans_num = []  # 选项作为列表被储存
        while len(ans_num) < 4:  # 四个选项，随机生成，相当于打乱选项顺序（题库中是正确答案总在第一个）
            j = rd.randint(1, 4)
            if j not in ans_num:
                ans_num.append(j)
        ques_box[ques_nu[len(ques_box)]] = ans_num  # 字典运用，字典的索引为之前随机生成的题号，对应内容为随机生成的选项编号
    return ques_box  # 返回生成的字典


def create_box():  # 将题目信息存于Ques类中并储存在列表中，方便后续取用
    # global ws, box, num_of_ch_ques, num_of_fill_ques, num_of_check
    cho_box = ques_num('choose', num_of_ch_ques + num_of_fill_ques)  # 选择填空题
    che_box = ques_num('check', num_of_check)  # 判断题
    key = list(cho_box.keys())  # 选择填空题的题号
    while len(box) < num_of_ch_ques + num_of_fill_ques + num_of_check:  # 三种题整合进一个列表
        if len(box) < num_of_ch_ques:  # 整合选择题
            ques = Ques(ws.cell(key[len(box)], 1).value, cho_box[key[len(box)]], key[len(box)])
        elif len(box) < num_of_fill_ques + num_of_ch_ques:  # 整合填空题
            ques = Ques(ws.cell(key[len(box)], 1).value, str(ws.cell(key[len(box)], 2).value), key[len(box)])
        else:  # 整合判断题
            ques = Ques(ws.cell(che_box[len(box) - num_of_fill_ques - num_of_ch_ques], 1).value,
                        int(ws.cell(che_box[len(box) - num_of_fill_ques - num_of_ch_ques], 2).value),
                        che_box[len(box) - num_of_fill_ques - num_of_ch_ques])
        box.append(ques)  # 本次选题的列表，元素为Ques类，方便后续代码操作


def begin():  # 组卷及重新组卷
    # global fra, cho_ans, box, num_of_ch_ques, num_of_fill_ques, tx_box
    global tx_box
    tx_box=[]  # 清空，保证不影响重组试卷之后的功能
    clr=0  # 颜色列表的标签，是每一行可以进行不同颜色的填充
    create_box()  # 生成整合后题目列表
    for i in box:  # 将题放入GUI
        text=tk.Text(fra[box.index(i)//5], font=15,bg=colors[clr%10])  # fra中有四个Frame()，代表有4页，每页5题
        text.insert('1.0', str(box.index(i) + 1) + '. ' + i.title)
        text.config(state=tk.DISABLED)  # 题干的放置
        text.place(relwidth=1,relheight=.12,relx=0,rely=(box.index(i)%5)*.2)
        clr += 1
        if box.index(i) < num_of_ch_ques:  # 对于选择题的选项可视化
            for j in i.answer:  # 4个选项同用一个放于cho_ans中的IntVar()，代表四个选项对应一个题（不能多选）
                rad = tk.Radiobutton(fra[box.index(i) // 5], font=15, selectcolor='#00ff00', width=15,bg=colors[clr%10],
                                     variable=cho_ans[box.index(i)], indicatoron=0, value=j,relief=tk.RIDGE,bd=10,
                                     text=check_title[i.answer.index(j)]+'. '+str(ws.cell(i.num, j + 1).value))
                rad.place(relwidth=0.24,relheight=.08,relx=i.answer.index(j)*0.25,rely=box.index(i)%5*.2+.12)
            clr += 1
        elif box.index(i) < num_of_ch_ques + num_of_fill_ques:  # 填空题答题框
            tx = tk.Text(fra[box.index(i) // 5], font=15, height=2,bg=colors[clr%10],)
            tx_box.append(tx)  # 后面要取用输入的内容来判断正误，所以放入列表整合
            tx.place(relwidth=1,relheight=.08,relx=0,rely=box.index(i)%5*.2+.12)
            clr += 1
        else:  # 判断题的选项
            rad1 = tk.Radiobutton(fra[box.index(i) // 5], font=15, selectcolor='#00ff00', width=15,relief=tk.RIDGE,bd=10,
                                  variable=che_ans[box.index(i) - num_of_fill_ques - num_of_ch_ques], value=1, text='对',
                                  indicatoron=0,bg=colors[clr%10])
            clr += 1
            # 正确选项的生成，下面为错误选项的生成
            rad2 = tk.Radiobutton(fra[box.index(i) // 5], font=15, selectcolor='#00ff00', width=15,relief=tk.RIDGE,bd=10,
                                  variable=che_ans[box.index(i) - num_of_fill_ques - num_of_ch_ques], value=-1,
                                  text='错', indicatoron=0,bg=colors[clr%10])
            clr += 1
            rad1.place(relwidth=.4,relheight=.08,relx=.05,rely=box.index(i)%5*.2+.12)
            rad2.place(relwidth=.4, relheight=.08, relx=.54, rely=box.index(i) % 5*.2+.12)


def correct_paper():  # 试卷判分
    # global cho_ans, che_ans, num_of_check
    score, wrong, yesornot = 0, [], ['错',None,'对']  # 总分和错误的题号
    for i in cho_ans:  # 选择题是否做对
        if i.get() == 1:
            score += 5
        else:
            wrong.append(str(cho_ans.index(i) + 1))  # 没做对将题号放入列表后续显示
    for i in range(num_of_ch_ques, num_of_fill_ques + num_of_ch_ques):  # 填空题正误
        if answer_process(tx_box[i - num_of_ch_ques].get(0.0, 'end')) == str(box[i].answer):
            score += 5
        else:
            wrong.append(str(i + 1))
    for i in che_ans:  # 判断题正误
        if i.get() == box[-num_of_check + che_ans.index(i)].answer:
            score += 5
        else:
            wrong.append(str(che_ans.index(i) + num_of_ch_ques + num_of_fill_ques + 1))
    history['state']=tk.NORMAL
    history.insert(tk.INSERT,'\n\n'+str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())))

    history.insert(tk.INSERT,'\n你的总分是：'+str(score))
    if len(wrong):
        history.insert(tk.INSERT,'\n你的错题有：'+'，'.join(wrong)+'\n试卷答案是：')
        for i in wrong:
            i = int(i)-1
            if i<num_of_ch_ques:
                history.insert(tk.INSERT,check_title[box[i].answer.index(1)]+'， ')
            elif i<num_of_ch_ques+num_of_fill_ques:
                history.insert(tk.INSERT, box[i].answer + '， ')
            else:
                history.insert(tk.INSERT, yesornot[box[i].answer+1] + '， ')
    history['state']=tk.DISABLED
    if score == 100:
        msb.showinfo('结果','你的总分是：'+str(score)+'\n恭喜你全答对了')
    else:
        msb.showinfo('结果', '你的总分是：' + str(score) + '\n你的错题有：'+'，'.join(wrong))


def correct():  # 改卷函数
    global can_update,can_correct
    if can_correct:
        boo = msb.askokcancel('提示', '你确定要提交试卷吗？')
        if boo:
            correct_paper()  # 改卷并打分生成历史记录
            can_update = True
            can_correct = False
    else:
        msb.showwarning('错误','本次你已交卷，请在历史记录中查看历次分数')


def remake_paper():  # 重新组卷函数
    global box,can_update,can_correct
    if not can_update:
        msb.showerror('错误','请先提交试卷后执行此操作')
    else:
        boo=msb.askokcancel('提示','确定要重新生成试卷？')
        if boo:
            box = []
            begin()  # 进行重新组卷
            can_update=False
            can_correct=True


def my_help():
    global img
    top = tk.Toplevel()
    tk.Label(top,image=img,text='其实没有帮助',compound='top').pack()


def pick(event):
    while 1:
        if event==1:
            im = Image.open('img2.gif')  # GIF图片流的迭代器
            can = can1
        else:
            im = Image.open('img3.gif')
            can = can2
        my_iter = ImageSequence.Iterator(im)  # frame就是gif的每一帧，转换一下格式就能显示了
        for frame in my_iter:
            pic=ImageTk.PhotoImage(frame)
            can.create_image(160,100, image=pic)
            time.sleep(0.1)  # 一秒10帧
            win.update_idletasks()  # 刷新
            win.update()


def change_page(event):
    global page_num
    if event.delta<0:  # 滚轴向下滚
        page_num+=1
    else:
        page_num-=1
    if page_num>4:
        page_num=0
    elif page_num<0:
        page_num=4
    note.select(fra[page_num])


def my_quit():
    boo = msb.askokcancel('提示','确定退出程序？')
    if boo:
        win.quit()


if __name__ == '__main__':
    win = tk.Tk()
    mu1 = tk.Menu(win)
    note = ttk.Notebook(win)
    wb = opl.load_workbook('test.xlsx')  # 加载题库
    ws = wb.active  # 激活题库，此后可调用.xlsx中的数据

    img = tk.PhotoImage(file='th.png')
    # 颜色列表
    colors=['lightcyan','powderblue','paleturquoise','lightblue','azure','lightcyan','honeydew','aquamarine','mediumaquamarine','lightseagreen']
    can_update=False  # 是否允许重新组卷的标志位
    can_correct=True  # 是否允许再次改卷的标志位
    check_title=['A','B','C','D']
    tx_box = []
    box = []  # 存储选好的题目的相关信息
    cho_ans = []  # 存储选择题按键的选值，用于后续判断对错,正确答案对应数值均为1
    che_ans = []  # 储存判断题按键选值
    fra = []  # 页面列表
    num_of_ch_ques = 10  # 选择题数目
    num_of_fill_ques = 5  # 填空题数目
    num_of_check = 5  # 判断题数目

    for ii in range(num_of_check):
        che_ans.append(tk.IntVar())
    for ii in range(num_of_ch_ques):
        cho_ans.append(tk.IntVar())
    for ii in range(5):
        fra.append(tk.Frame(win))  # 初始化变量
    for ii in fra:  # 将frame放入notebook形成标签页面效果
        note.add(ii, text=str(fra.index(ii) * 5 + 1) + '~' + str(fra.index(ii) * 5 + 5))
        if fra.index(ii)==4:
            note.add(ii, text='历史得分')
    win.title('基于tkinter的试卷组成')
    win.geometry('650x650')
    win.minsize(600, 650)
    can1 = tk.Canvas(fra[4],bg='white')  # 画布1
    can2 = tk.Canvas(fra[4], bg='white')  # 画布2
    can1.place(relwidth=.5,relheight=.3)
    can2.place(relwidth=.5, relheight=.3,relx=.5)
    history=tk.Text(fra[4],state=tk.DISABLED)
    scr = tk.Scrollbar(fra[4], command=history.yview)  # 滚动条
    history.config(yscrollcommand=scr.set)
    history.place(relwidth=.98,relheight=.7,rely=.3)
    scr.place(relx=.98,relwidth=.02,relheight=.7,rely=.3)
    mu1.add_command(label='提交',command=correct)
    mu1.add_command(label='重新组卷',command=remake_paper)
    mu1.add_command(label='退出',command=my_quit)
    mu1.add_separator()
    mu1.add_command(label='帮助',command=my_help)
    win.config(menu=mu1)
    note.place(relwidth=1, relheight=1)
    page_num = 0
    win.bind('<MouseWheel>',change_page)  # 鼠标滚轴滚动换页
    win.bind('<MouseWheel>',change_page)
    begin()

    can1.bind("<Double-Button-1>", lambda x:pick(1))  # 动画1
    can2.bind("<Double-Button-1>", lambda x:pick(2))  # 动画2

    tk.mainloop()
